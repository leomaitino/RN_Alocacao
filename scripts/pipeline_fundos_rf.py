"""
=============================================================================
ALPHA DESK — Pipeline de Dados (Renda Fixa)
=============================================================================
Gera os arquivos JSON que alimentam o dashboard de fundos de renda fixa.

DIFERENÇAS vs pipeline_fundos.py (multimercados):
  - NÃO filtra por CLASSIFICAÇÃO_CVM (alguns RF têm CVM=Multimercado).
  - Mapeia CLASSIFICAÇÃO_XP → 5 subgrupos: Liquidez D0, Crédito,
    Incentivadas, FIDCs, Internacionais (whitelist estrita).
  - Cada fundo tem benchmark próprio (CDI ou IMA-B 5) derivado da XP class.
  - Sharpe e Excesso ANUALIZADO (% a.a.) calculados vs benchmark de cada fundo.
  - Não tem corr_cdi_36m (não faz sentido em RF).
  - Adiciona suporte ao IMA-B 5 (--imab5 arquivo.xlsx). Append em
    benchmarks.json sem sobrescrever as chaves do MM.
  - Cache CVM próprio (data/cache_cvm_rf/) — não compartilha com MM,
    porque cache é filtrado por CNPJ.
  - Reutiliza gestoras.json e conteudo.json do MM (não cria templates novos).

OUTPUTS (pasta /data):
  fundos_rf.json     — base de fundos RF com subgrupo + benchmark + métricas
  cotas_rf.json      — séries normalizadas (base 100) só dos fundos RF
  meta_rf.json       — metadata + benchmark efetivo de Inflação
  benchmarks.json    — atualizado in-place adicionando IMA-B 5
  recomendados_rf.json — template (criado se não existir)

USO:
  python scripts/pipeline_fundos_rf.py --xp lista-fundos-rf.xlsx
  python scripts/pipeline_fundos_rf.py --xp lista-fundos-rf.xlsx --imab5 imab5.xlsx
=============================================================================
"""

import argparse
import json
import os
import re
import sys
import warnings
from datetime import datetime, date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# Silencia warnings cosméticos:
#  (a) RuntimeWarnings do NumPy/pandas em séries com NaN — calcular_volatilidade/
#      sharpe operam sobre tail() de retornos diários que naturalmente contêm
#      NaN nos primeiros dias após pct_change. Resultados finais estão corretos.
#  (b) UserWarning do pandas sobre dayfirst quando a coluna de data já vem em
#      formato ISO (YYYY-MM-DD HH:MM:SS) — comum no Excel ANBIMA do IMA-B 5.
#      `dayfirst=True` é redundante mas inofensivo nesse formato; o warning
#      é só ruído.
warnings.filterwarnings('ignore', category=RuntimeWarning,
                        message=r'invalid value encountered in (subtract|reduce)')
warnings.filterwarnings('ignore', category=UserWarning,
                        message=r'Parsing dates in .* format when dayfirst=True was specified')

# ─── Funções/objetos importados de pipeline_fundos.py (paridade com MM) ─────
# Em pipeline_fundos.py cada bloco compartilhado tem marcador [SHARED-RF].
# Alterar assinatura lá quebra o RF aqui — fazer em coordenação ou duplicar.
#
#   Helpers de parsing:    parse_pct, parse_float, tipo_investidor,
#                          _normalizar_cnpj, _to_serializable, _serie_para_dict
#   CVM (download):        baixar_cadastro_cvm, enriquecer_cadastro,
#                          baixar_informes_cvm
#   Cálculos:              calcular_retornos_diarios, calcular_acumulado
#   Métricas:              calcular_volatilidade, calcular_sharpe,
#                          calcular_sortino, calcular_drawdown_max,
#                          calcular_var_95, calcular_calmar
#   Benchmarks:            baixar_cdi, baixar_ipca, baixar_ihfa,
#                          _parsear_csv_ihfa, construir_ipca_mais_spread
#   Outros:                NumpyEncoder, log, DIAS_UTEIS_ANO,
#                          MIN_DIAS_CALCULO, IPCA_SPREADS
# ────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from pipeline_fundos import (  # noqa: E402
    NumpyEncoder,
    parse_pct, parse_float, tipo_investidor,
    baixar_cadastro_cvm, enriquecer_cadastro,
    _normalizar_cnpj, baixar_informes_cvm,
    calcular_retornos_diarios,
    calcular_volatilidade, calcular_sharpe, calcular_sortino,
    calcular_drawdown_max, calcular_var_95, calcular_calmar,
    baixar_cdi, baixar_ipca, baixar_ihfa, _parsear_csv_ihfa,
    construir_ipca_mais_spread,
    _serie_para_dict, calcular_acumulado,
    _to_serializable,
    log,
    DIAS_UTEIS_ANO, MIN_DIAS_CALCULO, IPCA_SPREADS,
)

# ---------------------------------------------------------------------------
# CONFIG — Mapeamento Subgrupo / Benchmark (Q3/Q4)
# ---------------------------------------------------------------------------

# Whitelist por subgrupo (B1: classes não listadas são excluídas com warning)
SUBGRUPO_LIQUIDEZ_D0 = {
    'Crédito Liquidez',
    'Referenciado DI Soberano',
}
SUBGRUPO_CREDITO = {
    'Crédito High Grade',
    'Crédito Híbrido',
    'Crédito High Yield',
    'Crédito Privado High Yield',
    'Renda Fixa Ativo',
    'Macro Baixa Vol',
    'Renda Fixa Inflação',
}
SUBGRUPO_INCENTIVADAS = {
    'Debêntures Incentivadas Hedgeado',
    'Debêntures Incentivadas Não Hedgeado',
    'Crédito Privado Debêntures Incentivadas',
    'Juros Ativo Incentivado',
}
SUBGRUPO_INTERNACIONAIS = {
    'Internacional Renda Fixa Hedgeado',
    'Internacional Renda Fixa Não Hedgeado',
}
SUBGRUPO_FIDCS_XP = {
    'Crédito Estruturado',
}

# Benchmark por CLASSIFICAÇÃO_XP (Q3 — derivado, não vem da planilha)
BENCHMARK_POR_XP = {
    'Crédito Liquidez':                            'CDI',
    'Referenciado DI Soberano':                    'CDI',
    'Crédito High Grade':                          'CDI',
    'Crédito Híbrido':                             'CDI',
    'Crédito High Yield':                          'CDI',
    'Crédito Privado High Yield':                  'CDI',
    'Renda Fixa Ativo':                            'CDI',
    'Macro Baixa Vol':                             'CDI',
    'Renda Fixa Inflação':                         'IMA-B 5',
    'Debêntures Incentivadas Hedgeado':            'CDI',
    'Debêntures Incentivadas Não Hedgeado':        'IMA-B 5',
    'Crédito Privado Debêntures Incentivadas':     'IMA-B 5',
    'Juros Ativo Incentivado':                     'CDI',
    'Crédito Estruturado':                         'CDI',
    'Internacional Renda Fixa Hedgeado':           'CDI',
    'Internacional Renda Fixa Não Hedgeado':       'CDI',
}

# Subgrupos que recebem ranking quantitativo (score). Os outros entram só como lista.
SUBGRUPOS_COM_RANKING = {'Crédito', 'Incentivadas', 'FIDCs'}

# Exclusões explícitas (logadas como informativas, não como warning)
EXCLUSOES_EXPLICITAS = {'Renda Fixa Pré'}

CVM_CLASS_FIDC = 'Classes de Cotas de Fundos FIDC'


# ---------------------------------------------------------------------------
# Subgrupo / Benchmark helpers
# ---------------------------------------------------------------------------

def _subgrupo_rf(class_xp: str, class_cvm: str) -> str | None:
    """
    Retorna o subgrupo RF do fundo, ou None se ele deve ser excluído.

    Ordem de prioridade (Q4):
      1. Internacional Renda Fixa * (A2: estrita, exclui outros 'Internacional *')
      2. FIDCs (CVM=FIDC ou XP='Crédito Estruturado')
      3. Incentivadas (XP contém 'Incentivad' ou é 'Juros Ativo Incentivado')
      4. Liquidez D0 (XP em ['Crédito Liquidez', 'Referenciado DI Soberano'])
      5. Crédito (B1: whitelist estrita; classes fora dela são excluídas)

    Exclusões silenciosas com retorno None:
      - 'Renda Fixa Pré'  → motivo explícito (logado como info)
      - Qualquer XP class fora de todas as whitelists → logado como warning
    """
    class_xp  = (class_xp  or '').strip()
    class_cvm = (class_cvm or '').strip()

    # Exclusão explícita (logged as info pelo caller)
    if class_xp in EXCLUSOES_EXPLICITAS:
        return None

    # 1. Internacional Renda Fixa — whitelist estrita (A2)
    if class_xp.startswith('Internacional Renda Fixa'):
        if class_xp in SUBGRUPO_INTERNACIONAIS:
            return 'Internacionais'
        return None  # variante desconhecida → exclui (logged as warning)

    # 2. FIDCs — CVM ou XP
    if class_cvm == CVM_CLASS_FIDC or class_xp in SUBGRUPO_FIDCS_XP:
        return 'FIDCs'

    # 3. Incentivadas — heurística + whitelist
    if 'Incentivad' in class_xp or class_xp == 'Juros Ativo Incentivado':
        if class_xp in SUBGRUPO_INCENTIVADAS:
            return 'Incentivadas'
        return None

    # 4. Liquidez D0 — exato
    if class_xp in SUBGRUPO_LIQUIDEZ_D0:
        return 'Liquidez D0'

    # 5. Crédito — whitelist estrita (B1)
    if class_xp in SUBGRUPO_CREDITO:
        return 'Crédito'

    return None  # XP class não mapeada


def _benchmark_rf(class_xp: str) -> str:
    """Retorna o benchmark do fundo derivado da CLASSIFICAÇÃO_XP."""
    class_xp = (class_xp or '').strip()
    return BENCHMARK_POR_XP.get(class_xp, 'CDI')


# ---------------------------------------------------------------------------
# ETAPA 1 — Leitura e parsing da planilha XP (RF)
# ---------------------------------------------------------------------------

def ler_planilha_xp_rf(caminho: str) -> pd.DataFrame:
    log.info(f"Lendo planilha XP (RF): {caminho}")

    import openpyxl
    wb_temp = openpyxl.load_workbook(caminho, read_only=True)
    abas = wb_temp.sheetnames
    wb_temp.close()
    aba_alvo = None
    for nome in ['Fundos', 'fundos', 'Planilha1', 'Sheet1']:
        if nome in abas:
            aba_alvo = nome
            break
    if aba_alvo is None:
        aba_alvo = abas[0]
    log.info(f"  Aba utilizada: '{aba_alvo}' (disponíveis: {abas})")
    wb_df = pd.read_excel(caminho, sheet_name=aba_alvo, dtype=str)
    wb_df.columns = [c.strip().upper().replace(' ', '_') for c in wb_df.columns]

    fundos = []
    contadores = {
        'total_linhas': 0,
        'incluidos': 0,
        'excluidos_pre': 0,
        'excluidos_nao_mapeado': 0,
        'por_subgrupo': {},
    }
    nao_mapeados = {}  # class_xp → contador (para resumo final)

    for _, row in wb_df.iterrows():
        contadores['total_linhas'] += 1
        class_cvm = (row.get('CLASSIFICAÇÃO_CVM') or '').strip()
        class_xp_raw = row.get('CLASSIFICAÇÃO_XP') or ''
        class_xp = str(class_xp_raw).strip() if class_xp_raw and str(class_xp_raw) != 'nan' else ''

        subgrupo = _subgrupo_rf(class_xp, class_cvm)

        if subgrupo is None:
            # Distingue motivo: exclusão explícita vs class desconhecida
            if class_xp in EXCLUSOES_EXPLICITAS:
                contadores['excluidos_pre'] += 1
            else:
                contadores['excluidos_nao_mapeado'] += 1
                nao_mapeados[class_xp or '(vazio)'] = nao_mapeados.get(class_xp or '(vazio)', 0) + 1
            continue

        benchmark = _benchmark_rf(class_xp)

        f = {
            # — Identidade
            "nome":               (row.get('NOME_FUNDO') or '').strip(),
            "cnpj":               (row.get('CNPJ_FUNDO') or '').strip(),  # mantém pontuado (Q5)
            "gestora":            (row.get('NOME_GESTORA') or '').strip(),
            # — Classificações
            "class_cvm":          class_cvm or 'Não classificado',
            "class_xp":           class_xp or 'Não classificado',
            "subgrupo":           subgrupo,
            "benchmark":          benchmark,
            "tem_ranking":        subgrupo in SUBGRUPOS_COM_RANKING,
            # — Acesso
            "captacao_aberta":    (row.get('CAPTAÇÃO') or '').strip() == 'Aberta',
            "tipo_investidor":    tipo_investidor(row.get('TIPO_INVESTIDOR')),
            "aplicacao_minima":   parse_float(row.get('APLICAÇÃO_INICIAL_MÍNIMA')),
            "movimentacao_minima":parse_float(row.get('MOVIMENTAÇÃO_MÍNIMA')),
            # — Liquidez
            "cotizacao_resgate":  (row.get('COTIZAÇÃO_RESGATE') or '').strip(),
            "periodo_cotizacao":  (row.get('PERÍODO_COTIZAÇÃO') or '').strip(),
            "liquidacao_resgate": (row.get('LIQUIDAÇÃO_RESGATE') or '').strip(),
            "periodo_liquidacao": (row.get('PERÍODO_LIQUIDAÇÃO') or '').strip(),
            # — Rentabilidades XP (já calculadas)
            "rent_dia":  parse_pct(row.get('RENTABILIDADE_DIA')),
            "rent_mes":  parse_pct(row.get('RENTABILIDADE_MÊS')),
            "rent_ano":  parse_pct(row.get('RENTABILIDADE_ANO')),
            "rent_12m":  parse_pct(row.get('RENTABILIDADE_12M')),
            "rent_24m":  parse_pct(row.get('RENTABILIDADE_24M')),
            "rent_36m":  parse_pct(row.get('RENTABILIDADE_36M')),
            # — Cota
            "valor_cota": parse_float(row.get('VALOR_COTA')),
            "data_cota":  (row.get('DATA_COTA') or '').strip(),
            # — Risco Gênio
            "risco_genio": int(float(row['RISCO_GÊNIO'])) if row.get('RISCO_GÊNIO') else None,
            # — Taxas
            "taxa_adm":     parse_float(row.get('TAXA_ADMINISTRAÇÃO')),
            "taxa_adm_max": parse_float(row.get('TAXA_ADMINISTRAÇÃO_MÁXIMA')),
            "taxa_perf":    parse_float(row.get('TAXA_PERFORMANCE')),
            # — Enriquecidos via CVM
            "pl":           None,
            "num_cotistas": None,
            "data_inicio":  None,
            # — Métricas calculadas
            "sharpe":            None,
            "sharpe_24m":        None,
            "sharpe_36m":        None,
            "sortino":           None,
            "sortino_24m":       None,
            "sortino_36m":       None,
            "volatilidade":      None,
            "volatilidade_24m":  None,
            "volatilidade_36m":  None,
            "drawdown_max":      None,
            "drawdown_max_36m":  None,
            "consistencia":      None,
            "consistencia_36m":  None,
            "var_95":            None,
            "calmar":            None,
            "meses_pos":         None,
            "total_meses":       None,
            "variacao_pl_12m":   None,
            "pl_12m_atras":      None,
            # — RF-específicas: excesso vs benchmark do fundo
            "excesso_12m":       None,
            "excesso_24m":       None,
            "excesso_36m":       None,
            # — Manuais (recomendados_rf.json)
            "longevidade_anos":  None,
            "recomendado":       False,
            "aprovado":          False,
        }
        fundos.append(f)
        contadores['incluidos'] += 1
        contadores['por_subgrupo'][subgrupo] = contadores['por_subgrupo'].get(subgrupo, 0) + 1

    df = pd.DataFrame(fundos)

    # Log resumo
    log.info(f"  → {contadores['total_linhas']} linhas na planilha")
    log.info(f"  → {contadores['incluidos']} fundos incluídos")
    log.info(f"  → {contadores['excluidos_pre']} excluídos: 'Renda Fixa Pré' (não recomendado)")
    log.info(f"  → {contadores['excluidos_nao_mapeado']} excluídos: CLASSIFICAÇÃO_XP não mapeada")
    if nao_mapeados:
        log.warning("  Classes XP não mapeadas (excluídas):")
        for cls, n in sorted(nao_mapeados.items(), key=lambda kv: -kv[1]):
            log.warning(f"    × {cls}: {n} fundo(s)")
    log.info("  Distribuição por subgrupo:")
    for sg in ['Liquidez D0', 'Crédito', 'Incentivadas', 'FIDCs', 'Internacionais']:
        n = contadores['por_subgrupo'].get(sg, 0)
        ranking = ' [com ranking]' if sg in SUBGRUPOS_COM_RANKING else ' [só lista]'
        log.info(f"    • {sg}: {n}{ranking}")
    return df


# ---------------------------------------------------------------------------
# ETAPA 2 — Métricas (RF) — primeira passagem (sem benchmark)
# ---------------------------------------------------------------------------

def enriquecer_metricas_rf(df_fundos: pd.DataFrame, df_cotas: pd.DataFrame):
    """
    Calcula métricas que NÃO dependem do benchmark do fundo:
    PL, número de cotistas, var_PL_12M, vol, drawdown, var_95, consistência.

    Retorna (df_atualizado, dict_metricas_temp). O dict guarda séries de
    retornos/cotas para a segunda passagem (Sharpe + Excesso vs benchmark).
    """
    if df_cotas.empty:
        log.warning("  Sem dados de cotas CVM — métricas quant não calculadas")
        return df_fundos, {}

    log.info("Calculando métricas quantitativas via cotas CVM...")

    df_cotas = df_cotas.copy()
    df_cotas['CNPJ_NORM'] = df_cotas['CNPJ_FUNDO'].str.replace(r'[./-]', '', regex=True)

    metricas = {}
    for cnpj_norm, grupo in df_cotas.groupby('CNPJ_NORM'):
        grupo = (grupo.sort_values('DT_COMPTC')
                       .drop_duplicates('DT_COMPTC', keep='last')
                       .set_index('DT_COMPTC'))
        cotas = grupo['VL_QUOTA'].dropna()
        if len(cotas) < 5:
            continue

        ret = calcular_retornos_diarios(cotas)

        # PL e cotistas mais recentes
        pl_recente = grupo['VL_PATRIM_LIQ'].dropna().iloc[-1] if not grupo['VL_PATRIM_LIQ'].dropna().empty else None
        cotistas_recente = grupo['NR_COTST'].dropna().iloc[-1] if not grupo['NR_COTST'].dropna().empty else None

        # Variação PL (12M)
        pl_serie = grupo['VL_PATRIM_LIQ'].dropna()
        pl_atual = float(pl_recente) / 1e6 if pl_recente else None
        pl_12m_atras = None
        variacao_pl_12m = None
        if len(pl_serie) >= DIAS_UTEIS_ANO:
            pl_12m_atras = float(pl_serie.iloc[-DIAS_UTEIS_ANO]) / 1e6
            # Guard: pl_atual pode ser None se PL recente é 0/None mas o fundo
            # tinha PL 12M atrás. Edge case real (fundo que foi a quase-zero).
            if pl_atual is not None and pl_12m_atras and pl_12m_atras > 0:
                variacao_pl_12m = round((pl_atual - pl_12m_atras) / pl_12m_atras, 6)

        # Consistência (% meses positivos) — geral e 36M
        ret_mensal = (1 + ret).resample('ME').prod() - 1
        meses_pos = int((ret_mensal > 0).sum())
        total_meses = len(ret_mensal)
        consistencia = round(meses_pos / total_meses, 4) if total_meses > 0 else None

        JANELA_36M = DIAS_UTEIS_ANO * 3
        ret_36m   = ret.tail(JANELA_36M)
        cotas_36m = cotas.tail(JANELA_36M)

        ret_mensal_36m  = (1 + ret_36m).resample('ME').prod() - 1
        meses_pos_36m   = int((ret_mensal_36m > 0).sum())
        total_meses_36m = len(ret_mensal_36m)
        consistencia_36m = round(meses_pos_36m / total_meses_36m, 4) if total_meses_36m > 0 else None

        metricas[cnpj_norm] = {
            'pl':               round(pl_atual, 2) if pl_atual else None,
            'pl_12m_atras':     round(pl_12m_atras, 2) if pl_12m_atras else None,
            'variacao_pl_12m':  variacao_pl_12m,
            'num_cotistas':     int(cotistas_recente) if cotistas_recente else None,
            'volatilidade':     round(calcular_volatilidade(ret) or 0, 6),
            'volatilidade_24m': round(calcular_volatilidade(ret, janela=DIAS_UTEIS_ANO * 2) or 0, 6),
            'volatilidade_36m': round(calcular_volatilidade(ret_36m) or 0, 6),
            'drawdown_max':     round(calcular_drawdown_max(cotas) or 0, 6),
            'drawdown_max_36m': round(calcular_drawdown_max(cotas_36m) or 0, 6),
            'var_95':           round(calcular_var_95(ret) or 0, 6),
            'consistencia':     consistencia,
            'consistencia_36m': consistencia_36m,
            'meses_pos':        meses_pos,
            'total_meses':      total_meses,
            '_ret':             ret,
            '_cotas':           cotas,
        }

    log.info(f"  → Métricas (1ª passagem) calculadas para {len(metricas)} fundos")

    def aplicar(row):
        cnpj_n = _normalizar_cnpj(row['cnpj'])
        m = metricas.get(cnpj_n, {})
        for campo in ['pl', 'pl_12m_atras', 'variacao_pl_12m', 'num_cotistas',
                      'volatilidade', 'volatilidade_24m', 'volatilidade_36m',
                      'drawdown_max', 'drawdown_max_36m', 'var_95',
                      'consistencia', 'consistencia_36m',
                      'meses_pos', 'total_meses']:
            if campo in m:
                row[campo] = m[campo]
        row['calmar'] = calcular_calmar(row.get('rent_12m'), row.get('drawdown_max'))
        cotas = m.get('_cotas')
        if cotas is not None and len(cotas) > 0:
            row['longevidade_anos'] = round(len(cotas) / DIAS_UTEIS_ANO, 2)
        return row

    df_fundos = df_fundos.apply(aplicar, axis=1)
    return df_fundos, metricas


# ---------------------------------------------------------------------------
# ETAPA 3 — Excesso e Sharpe vs benchmark do fundo (segunda passagem)
# ---------------------------------------------------------------------------

def calcular_excesso_anualizado(cotas: pd.Series,
                                 retornos_benchmark: pd.Series,
                                 dias: int) -> float | None:
    """
    Excesso ANUALIZADO (% a.a.) sobre o benchmark do fundo na janela de `dias`
    dias úteis.

    Fórmula:
        E_cum = ret_acum_fundo - ret_acum_benchmark  (cumulativo na janela)
        excesso_anualizado = (1 + E_cum) ** (1 / n_anos) - 1
    com n_anos = dias / DIAS_UTEIS_ANO.

    Comportamento por janela:
        12M (n_anos=1): cumulativo == anualizado
        24M (n_anos=2): (1+E_cum)**0.5     - 1
        36M (n_anos=3): (1+E_cum)**(1/3)   - 1

    Por que anualizar: indústria de RF fala em "% a.a. vs benchmark";
    cumulativo distorce comparação entre janelas (24M parece sempre "melhor"
    que 12M em cumulativo só pelo prazo).

    Alinha pelas datas em que o fundo tem cota; exige cobertura mínima de 80%
    do benchmark dentro dessa janela.
    """
    if cotas is None or len(cotas) < dias + 1:
        return None
    # Retornos diários do fundo na janela. tail(dias+1) → pct_change() → dias retornos.
    ret_fundo = cotas.tail(dias + 1).pct_change().dropna()
    if len(ret_fundo) < dias:
        return None

    # Apenas datas em que AMBOS têm dado — apples-to-apples sobre o mesmo conjunto.
    bench_alinhado = retornos_benchmark.reindex(ret_fundo.index).dropna()
    if len(bench_alinhado) < dias * 0.8:
        return None
    ret_fundo_alinhado = ret_fundo.reindex(bench_alinhado.index)

    ret_fundo_cum = float((1 + ret_fundo_alinhado).prod() - 1)
    ret_bench_cum = float((1 + bench_alinhado).prod() - 1)
    excesso_cum = ret_fundo_cum - ret_bench_cum

    n_anos = dias / DIAS_UTEIS_ANO
    base = 1 + excesso_cum
    if base <= 0:
        # E_cum <= -100% — não dá pra anualizar via potência fracionária
        return None
    return round(base ** (1 / n_anos) - 1, 6)


def segunda_passagem_sharpe_excesso_rf(df_fundos: pd.DataFrame,
                                        benchmark_series_map: dict,
                                        metricas: dict) -> pd.DataFrame:
    """
    Para cada fundo, calcula Sharpe e Sortino (12M/24M/36M) e Excesso (12/24/36M)
    usando a SÉRIE DE RETORNOS DIÁRIOS DO BENCHMARK PRÓPRIO DO FUNDO
    (lookup em benchmark_series_map['CDI'] ou ['IMA-B 5']).

    benchmark_series_map: {'CDI': pd.Series, 'IMA-B 5': pd.Series}
    """
    if not metricas or not benchmark_series_map:
        return df_fundos

    serie_cdi = benchmark_series_map.get('CDI', pd.Series(dtype=float))

    def aplicar(row):
        cnpj_n = _normalizar_cnpj(row['cnpj'])
        m = metricas.get(cnpj_n, {})
        ret = m.get('_ret')
        cotas = m.get('_cotas')
        if ret is None or ret.empty:
            return row

        bench_name = row.get('benchmark', 'CDI')
        bench_serie = benchmark_series_map.get(bench_name)
        if bench_serie is None or bench_serie.empty:
            # fallback duro: CDI (não deveria acontecer, mas evita travar a passagem)
            bench_serie = serie_cdi

        J1 = DIAS_UTEIS_ANO
        J2 = DIAS_UTEIS_ANO * 2
        J3 = DIAS_UTEIS_ANO * 3

        # Sharpe e Sortino vs benchmark do fundo
        row['sharpe']      = round(calcular_sharpe(ret, bench_serie, janela=J1) or 0, 4)
        row['sharpe_24m']  = round(calcular_sharpe(ret, bench_serie, janela=J2) or 0, 4)
        row['sharpe_36m']  = round(calcular_sharpe(ret, bench_serie, janela=J3) or 0, 4)
        row['sortino']     = round(calcular_sortino(ret, bench_serie, janela=J1) or 0, 4)
        row['sortino_24m'] = round(calcular_sortino(ret, bench_serie, janela=J2) or 0, 4)
        row['sortino_36m'] = round(calcular_sortino(ret, bench_serie, janela=J3) or 0, 4)

        # Excesso anualizado (% a.a.) vs benchmark do fundo
        row['excesso_12m'] = calcular_excesso_anualizado(cotas, bench_serie, J1)
        row['excesso_24m'] = calcular_excesso_anualizado(cotas, bench_serie, J2)
        row['excesso_36m'] = calcular_excesso_anualizado(cotas, bench_serie, J3)
        return row

    df_fundos = df_fundos.apply(aplicar, axis=1)
    return df_fundos


# ---------------------------------------------------------------------------
# ETAPA 4 — IMA-B 5 (download/parse Excel ANBIMA)
# ---------------------------------------------------------------------------

def _parsear_excel_imab5(caminho: str, sheet_hint: str | None = None) -> pd.Series:
    """
    Parser flexível do Excel ANBIMA do IMA Histórico.

    Estratégia para escolher a aba:
      0. Se `sheet_hint` (vindo de --imab5-sheet) for passado, tenta usar
         exatamente essa aba; se não existir, falha e retorna série vazia
         (não cai pra heurística — explícito vence implícito).
      1. Procura aba cujo nome contém 'IMA-B 5' mas NÃO termina com '+' nem
         tem '+' depois do '5' (para distinguir IMA-B 5 vs IMA-B 5+).
      2. Fallback: se o arquivo tem só 1 aba, usa ela (caso usuário tenha
         exportado só o IMA-B 5 num Excel mono-aba).

    Detecta colunas de data e de número-índice ou variação diária.
    Converte para série de retornos diários (decimal).
    """
    import openpyxl
    wb = openpyxl.load_workbook(caminho, read_only=True)
    abas = wb.sheetnames
    wb.close()

    aba_alvo = None
    if sheet_hint:
        # Aba explícita via --imab5-sheet
        if sheet_hint in abas:
            aba_alvo = sheet_hint
            log.info(f"  IMA-B 5: usando aba forçada via --imab5-sheet: '{aba_alvo}'")
        else:
            log.warning(f"  IMA-B 5: aba '{sheet_hint}' (--imab5-sheet) não existe. "
                        f"Abas disponíveis: {abas}")
            return pd.Series(dtype=float)
    else:
        # Heurística: procura "IMA-B 5", excluindo "IMA-B 5+"
        candidatas = []
        for nome in abas:
            nome_norm = nome.strip().upper().replace(' ', '')
            if 'IMA-B5' in nome_norm and '5+' not in nome_norm and '5MAIS' not in nome_norm:
                candidatas.append(nome)
        if candidatas:
            aba_alvo = candidatas[0]
        else:
            # fallback: aceita usuário ter mandado planilha de aba única só do IMA-B 5
            if len(abas) == 1:
                aba_alvo = abas[0]
                log.warning(f"  IMA-B 5: nenhuma aba 'IMA-B 5' encontrada; usando única aba "
                            f"'{aba_alvo}'. Para forçar use --imab5-sheet NOME.")
            else:
                log.warning(f"  IMA-B 5: nenhuma aba reconhecida. Abas: {abas}. "
                            f"Use --imab5-sheet NOME para forçar.")
                return pd.Series(dtype=float)

    log.info(f"  IMA-B 5: lendo aba '{aba_alvo}'")

    # Tenta múltiplos header_rows porque ANBIMA varia o layout
    for header_row in [0, 1, 2, 3]:
        try:
            df = pd.read_excel(caminho, sheet_name=aba_alvo, header=header_row, dtype=str)
            if df.shape[1] < 2:
                continue
            df.columns = [str(c).strip() for c in df.columns]

            col_data = next(
                (c for c in df.columns if any(p in str(c).lower() for p in ['data', 'dt_'])),
                None,
            )
            # Prefere "Variação Diária" sobre "Número Índice" — mais direto
            col_var = next(
                (c for c in df.columns if 'variação' in str(c).lower() and 'diária' in str(c).lower()),
                None,
            )
            col_idx = next(
                (c for c in df.columns if any(p in str(c).lower() for p in ['número', 'numero']) and 'índice' in str(c).lower() or 'indice' in str(c).lower()),
                None,
            )

            if not col_data or (not col_var and not col_idx):
                continue

            # Excel pode entregar a coluna já como datetime — pular conversão
            # para silenciar UserWarning de dayfirst em valores datetime nativos.
            if not pd.api.types.is_datetime64_any_dtype(df[col_data]):
                df[col_data] = pd.to_datetime(df[col_data], dayfirst=True, errors='coerce')
            df = df.dropna(subset=[col_data]).set_index(col_data).sort_index()

            if col_var:
                # Variação diária em % → decimal
                ret = pd.to_numeric(
                    df[col_var].astype(str).str.replace(',', '.', regex=False).str.replace('%', '', regex=False),
                    errors='coerce',
                ).dropna() / 100.0
            else:
                # Número-índice → calcula retorno diário
                idx = pd.to_numeric(
                    df[col_idx].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False),
                    errors='coerce',
                ).dropna()
                ret = idx.pct_change().dropna()

            if len(ret) > 100:
                return ret
        except Exception:
            continue

    log.warning(f"  IMA-B 5: não foi possível parsear a aba '{aba_alvo}'")
    return pd.Series(dtype=float)


def baixar_imab5(caminho_local: str | None,
                  sheet_hint: str | None = None) -> pd.Series:
    """
    Carrega série diária do IMA-B 5 a partir de arquivo Excel local.

    Sem download automático: ANBIMA bloqueia bots. Usuário deve baixar de
    https://www.anbima.com.br/pt_br/informar/estatisticas/precos-e-indices/ima-historico.htm
    e passar via --imab5. Opcionalmente, --imab5-sheet força o nome da aba.
    """
    if not caminho_local:
        log.warning("  IMA-B 5: --imab5 não fornecido")
        return pd.Series(dtype=float)
    if not os.path.exists(caminho_local):
        log.warning(f"  IMA-B 5: arquivo não encontrado: {caminho_local}")
        return pd.Series(dtype=float)

    log.info(f"Carregando IMA-B 5 do arquivo: {caminho_local}")
    serie = _parsear_excel_imab5(caminho_local, sheet_hint=sheet_hint)
    if not serie.empty:
        log.info(f"  → IMA-B 5: {len(serie)} dias ({serie.index[0].date()} a {serie.index[-1].date()})")
    return serie


# ---------------------------------------------------------------------------
# ETAPA 5 — Benchmarks (merge não-destrutivo) + IMA-B (longo)
# ---------------------------------------------------------------------------

def carregar_benchmarks_existente(output_dir: Path) -> dict:
    path = output_dir / 'benchmarks.json'
    if not path.exists():
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        log.warning(f"  Não foi possível ler benchmarks.json existente: {e}")
        return {}


def carregar_imab_existente(benchmarks_atuais: dict) -> pd.Series:
    """Reaproveita IMA-B do benchmarks.json se MM já tiver baixado."""
    if 'IMA-B' not in benchmarks_atuais:
        return pd.Series(dtype=float)
    rd = benchmarks_atuais['IMA-B'].get('retornos_diarios') or {}
    if not rd:
        return pd.Series(dtype=float)
    serie = pd.Series(
        {pd.Timestamp(k): v for k, v in rd.items()},
        dtype=float,
    ).sort_index()
    return serie


def montar_e_atualizar_benchmarks(benchmarks_atuais: dict,
                                   serie_cdi: pd.Series,
                                   serie_ipca: pd.Series,
                                   serie_ihfa: pd.Series,
                                   serie_imab_longo: pd.Series,
                                   serie_imab5: pd.Series) -> tuple[dict, str, bool]:
    """
    Atualiza benchmarks.json com IMA-B 5 (real ou fallback). Política de
    sobrescrita por chave (corrige bug de defasagem — antes preservava tudo):

      - CDI:      SOBRESCREVE quando serie_cdi (BCB API) está fresca.
                  Side effect positivo: MM em produção também se beneficia.
      - IPCA+X:   SOBRESCREVE quando serie_ipca está fresca.
      - IMA-B 5:  sempre escreve (real ou fallback IMA-B longo).
      - IMA-B (longo): PRESERVA — RF não tem fonte fresca pra essa série
                  (sem arquivo Anbima local). Ver BACKLOG #10.
      - IHFA:     PRESERVA — Anbima bloqueia bot, download falha. Ver BACKLOG #11.

    Retorna (benchmarks_atualizados, benchmark_inflacao_efetivo, fallback_usado).
    """
    benchmarks = dict(benchmarks_atuais)  # cópia rasa — preserva tudo do MM por default

    # CDI: sobrescreve com dados frescos da rodada atual
    if not serie_cdi.empty:
        acum = calcular_acumulado(serie_cdi)
        benchmarks['CDI'] = {
            'nome': 'CDI', 'tipo': 'taxa',
            'retornos_diarios': _serie_para_dict(serie_cdi),
            'acumulado': _serie_para_dict(acum),
        }
    # IPCA+spreads: sobrescreve quando há IPCA + CDI frescos
    for label, spread in IPCA_SPREADS.items():
        s = construir_ipca_mais_spread(serie_ipca, spread, serie_cdi)
        if not s.empty:
            benchmarks[label] = {
                'nome': label, 'tipo': 'ipca_spread', 'spread_anual': spread,
                'retornos_diarios': _serie_para_dict(s),
                'acumulado': _serie_para_dict(calcular_acumulado(s)),
            }
    # IHFA: PRESERVA — RF não consegue baixar (Anbima bloqueia bot). BACKLOG #11.
    if 'IHFA' not in benchmarks and not serie_ihfa.empty:
        acum = calcular_acumulado(serie_ihfa)
        benchmarks['IHFA'] = {
            'nome': 'IHFA', 'tipo': 'indice',
            'retornos_diarios': _serie_para_dict(serie_ihfa),
            'acumulado': _serie_para_dict(acum),
        }
    # IMA-B (longo): PRESERVA — RF não tem fonte fresca, depende do MM. BACKLOG #10.
    if 'IMA-B' not in benchmarks and not serie_imab_longo.empty:
        acum = calcular_acumulado(serie_imab_longo)
        benchmarks['IMA-B'] = {
            'nome': 'IMA-B', 'tipo': 'indice',
            'retornos_diarios': _serie_para_dict(serie_imab_longo),
            'acumulado': _serie_para_dict(acum),
        }

    # IMA-B 5 — sempre escreve (real ou fallback) para o front sempre encontrar
    fallback_usado = False
    if not serie_imab5.empty:
        acum = calcular_acumulado(serie_imab5)
        benchmarks['IMA-B 5'] = {
            'nome': 'IMA-B 5', 'tipo': 'indice', 'fallback': False, 'origem': 'ANBIMA',
            'retornos_diarios': _serie_para_dict(serie_imab5),
            'acumulado': _serie_para_dict(acum),
        }
        infl_efetivo = 'IMA-B 5'
    elif not serie_imab_longo.empty:
        # Fallback: copia série do IMA-B longo
        acum = calcular_acumulado(serie_imab_longo)
        benchmarks['IMA-B 5'] = {
            'nome': 'IMA-B 5', 'tipo': 'indice', 'fallback': True,
            'origem': 'fallback IMA-B longo (IMA-B 5 indisponível)',
            'retornos_diarios': _serie_para_dict(serie_imab_longo),
            'acumulado': _serie_para_dict(acum),
        }
        infl_efetivo = 'IMA-B (fallback)'
        fallback_usado = True
        log.warning("  IMA-B 5 efetivo: FALLBACK para IMA-B longo")
    else:
        infl_efetivo = '(indisponível)'
        log.warning("  IMA-B 5 efetivo: NENHUM (nem IMA-B 5 nem IMA-B longo disponíveis)")

    log.info(f"  → benchmarks.json: {len(benchmarks)} chaves ({list(benchmarks.keys())})")
    return benchmarks, infl_efetivo, fallback_usado


# ---------------------------------------------------------------------------
# ETAPA 6 — Templates manuais (apenas recomendados_rf.json)
# ---------------------------------------------------------------------------

TEMPLATE_RECOMENDADOS_RF = {
    "_instrucoes": ("Adicione CNPJs (sem pontuação) dos fundos RF recomendados/aprovados. "
                    "Apenas Crédito, Incentivadas e FIDCs aceitam recomendação."),
    "_ultima_revisao": "",
    "recomendados": [],
    "aprovados": [],
}


def criar_templates_rf(output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / 'recomendados_rf.json'
    if path.exists():
        log.info(f"  → Arquivo manual já existe: recomendados_rf.json (mantido)")
        return
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(TEMPLATE_RECOMENDADOS_RF, f, ensure_ascii=False, indent=2)
    log.info(f"  → Template criado: recomendados_rf.json (preencha manualmente)")


def aplicar_recomendados_rf(df_fundos: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    path = output_dir / 'recomendados_rf.json'
    if not path.exists():
        return df_fundos
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        cnpjs_rec = {_normalizar_cnpj(str(c)) for c in data.get('recomendados', [])}
        cnpjs_apr = {_normalizar_cnpj(str(c)) for c in data.get('aprovados', [])}
        df_fundos['recomendado'] = df_fundos['cnpj'].apply(_normalizar_cnpj).isin(cnpjs_rec)
        df_fundos['aprovado']    = df_fundos['cnpj'].apply(_normalizar_cnpj).isin(cnpjs_apr)
        # Liquidez D0 e Internacionais não aceitam recomendado (regra de negócio)
        sem_rec = ~df_fundos['tem_ranking']
        rec_invalidos = (df_fundos['recomendado'] & sem_rec).sum()
        if rec_invalidos:
            log.warning(f"  → {rec_invalidos} recomendados em subgrupos sem ranking — desmarcados")
            df_fundos.loc[sem_rec, 'recomendado'] = False
            df_fundos.loc[sem_rec, 'aprovado']    = False
        log.info(f"  → {len(cnpjs_rec)} recomendados | {len(cnpjs_apr)} aprovados aplicados")
    except Exception as e:
        log.warning(f"  ✗ Falha ao ler recomendados_rf.json: {e}")
    return df_fundos


# ---------------------------------------------------------------------------
# ETAPA 7 — Outputs RF
# ---------------------------------------------------------------------------

def gerar_fundos_rf_json(df_fundos: pd.DataFrame) -> list[dict]:
    """
    Converte df_fundos em lista de dicts JSON-safe.

    Sanitiza inf/NaN → None: o spec do JSON não permite Infinity/NaN; o
    JSON.parse do JS rejeita com SyntaxError. Algumas funções de cálculo
    (calcular_sortino especialmente, e excesso anualizado em edge cases)
    retornam inf quando o downside_dev é minúsculo mas não-zero — comum em
    fundos D0 de vol baixíssima. Ver BACKLOG_RF.md #6.
    """
    colunas_excluir = {'_ret', '_cotas'}
    registros = []
    nao_finitos = {}  # campo -> [nomes] para log
    for _, row in df_fundos.iterrows():
        d = {}
        nome_row = row.get('nome', '?')
        for k, v in row.items():
            if k in colunas_excluir:
                continue
            v_serial = _to_serializable(v)
            if isinstance(v_serial, float) and not np.isfinite(v_serial):
                nao_finitos.setdefault(k, []).append(nome_row)
                v_serial = None
            d[k] = v_serial
        registros.append(d)
    if nao_finitos:
        log.warning("  [JSON_SANITIZE] valores inf/nan substituídos por None:")
        for campo in sorted(nao_finitos):
            nomes = nao_finitos[campo]
            log.warning(f"  [JSON_SANITIZE]   {campo}: {len(nomes)} fundo(s)")
            for n in nomes[:5]:
                log.warning(f"  [JSON_SANITIZE]     · {n}")
    return registros


def gerar_cotas_rf_json(fundos_list: list[dict],
                         cache_cvm_dir: Path,
                         df_cotas: pd.DataFrame = None) -> dict:
    """
    Monta cotas_rf.json APENAS para os CNPJs RF, com cota normalizada base 100.

    Fontes (em ordem de preferência):
      1. df_cotas em memória (passado por salvar_outputs_rf) — inclui o mês
         CORRENTE, que NÃO está nos parquets de cache (eh_mes_atual=True no
         baixar_informes_cvm não cacheia o mês atual por design). Ver
         BACKLOG_RF #9: mesmo bug existe no MM.
      2. Fallback para os parquets cacheados — defasagem de até 1 mês em
         relação à data atual (mês corrente ausente).

    Mesma lógica de dedup-por-continuidade do MM.
    """
    cotas_dict = {}

    df_c = None
    if df_cotas is not None and not df_cotas.empty:
        df_c = df_cotas.copy()
        log.info(f"  cotas_rf.json: fonte = df em memória "
                 f"({len(df_c)} registros, inclui mês corrente)")
    elif cache_cvm_dir.exists():
        import glob
        parquet_files = sorted(glob.glob(str(cache_cvm_dir / '*.parquet')))
        if parquet_files:
            try:
                frames = [pd.read_parquet(pf) for pf in parquet_files]
                df_c = pd.concat(frames, ignore_index=True)
                log.warning(f"  cotas_rf.json: fonte = fallback parquets "
                            f"({len(parquet_files)} meses) — mês corrente AUSENTE, "
                            f"gráficos vão ficar defasados")
            except Exception as e:
                log.warning(f"  ✗ Falha lendo parquets: {e}")
    if df_c is None or df_c.empty:
        log.warning(f"  Nenhuma fonte de cotas disponível — cotas_rf.json vazio")
        return cotas_dict

    try:
        df_c['DT_COMPTC'] = pd.to_datetime(df_c['DT_COMPTC'])
        # Detecta coluna de quota
        quota_col = next(
            (col for col in df_c.columns if 'QUOTA' in col.upper() or 'VL_COTA' in col.upper()),
            None,
        )
        if not quota_col:
            log.warning(f"  cotas_rf.json: coluna VL_QUOTA não encontrada. Cols: {list(df_c.columns)}")
            return cotas_dict
        if quota_col != 'VL_QUOTA':
            df_c = df_c.rename(columns={quota_col: 'VL_QUOTA'})
        if df_c['VL_QUOTA'].dtype == object:
            df_c['VL_QUOTA'] = pd.to_numeric(
                df_c['VL_QUOTA'].str.replace(',', '.', regex=False), errors='coerce')
        else:
            df_c['VL_QUOTA'] = pd.to_numeric(df_c['VL_QUOTA'], errors='coerce')

        df_c = df_c.sort_values(['CNPJ_NORM', 'DT_COMPTC'])
        cnpj_map = {re.sub(r'[./-]', '', f['cnpj']): f['cnpj'] for f in fundos_list}
        cotas_sanitizadas = []  # [(cnpj, motivo)] para log

        for cnpj_norm, grp in df_c.groupby('CNPJ_NORM'):
            cnpj_orig = cnpj_map.get(cnpj_norm)
            if not cnpj_orig:
                continue
            grp = grp.sort_values('DT_COMPTC')
            indices_manter = []
            prev_val = None
            for dt, dia_rows in grp.groupby('DT_COMPTC', sort=True):
                if len(dia_rows) == 1:
                    idx = dia_rows.index[0]
                    prev_val = dia_rows.iloc[0]['VL_QUOTA']
                else:
                    if prev_val is not None and pd.notna(prev_val):
                        diffs = (dia_rows['VL_QUOTA'] - prev_val).abs()
                        idx = diffs.idxmin()
                    else:
                        idx = dia_rows['VL_QUOTA'].idxmax()
                    prev_val = grp.loc[idx, 'VL_QUOTA']
                indices_manter.append(idx)
            grp = grp.loc[indices_manter].set_index('DT_COMPTC')
            cotas = grp['VL_QUOTA'].dropna()
            if len(cotas) < 2:
                continue
            base = cotas.iloc[0]
            # Guard: base ZERO ou não-finito → série inteira viraria inf/NaN
            # (raro mas acontece em FIDCs novos com cota inicial 0).
            if not np.isfinite(base) or base == 0:
                cotas_sanitizadas.append((cnpj_orig, 'base inválida'))
                continue
            norm = (cotas / base * 100).round(4)
            # Filtra qualquer ponto não-finito que tenha sobrado (defensivo)
            datas_norm = [d.strftime('%Y-%m-%d') for d in norm.index]
            valores_norm = norm.tolist()
            par_validos = [(d, v) for d, v in zip(datas_norm, valores_norm)
                           if isinstance(v, (int, float)) and np.isfinite(v)]
            n_descartados = len(valores_norm) - len(par_validos)
            if n_descartados > 0:
                cotas_sanitizadas.append((cnpj_orig, f'{n_descartados} pontos inf/nan'))
            if len(par_validos) < 2:
                continue
            datas_f, valores_f = zip(*par_validos)
            cotas_dict[cnpj_orig] = {
                'datas':   list(datas_f),
                'valores': list(valores_f),
            }
        if cotas_sanitizadas:
            log.warning(f"  [COTAS_SANITIZE] {len(cotas_sanitizadas)} série(s) com inf/nan tratadas:")
            for cnpj, motivo in cotas_sanitizadas[:10]:
                log.warning(f"  [COTAS_SANITIZE]   · {cnpj}: {motivo}")
        log.info(f"  → cotas_rf.json: {len(cotas_dict)} séries")
    except Exception as e:
        log.warning(f"  ✗ cotas_rf.json não gerado: {e}")
    return cotas_dict


def salvar_outputs_rf(df_fundos: pd.DataFrame,
                      benchmarks: dict,
                      output_dir: Path,
                      cache_cvm_dir: Path,
                      benchmark_inflacao_efetivo: str,
                      fallback_usado: bool,
                      df_cotas: pd.DataFrame = None):
    output_dir.mkdir(parents=True, exist_ok=True)

    # fundos_rf.json — filtra fundos sem volatilidade (sem dados CVM confiáveis).
    # Exceções preservadas:
    #   - Liquidez D0: vol baixíssima é inerente, não falta de dados.
    #   - FIDCs: registros recentes (2024-2025) sem informe diário CVM ainda
    #     publicado — ver BACKLOG_RF #3. Recomendamos esses fundos ativamente,
    #     então precisam estar na lista (front mostra sem score).
    # Cada fundo carrega flag `sem_dados_cvm: bool` (futuro-proof, para qualquer
    # subgrupo) — front usa pra decidir se exibe métricas calculadas ou só
    # rentabilidades da planilha XP.
    fundos_list_full = gerar_fundos_rf_json(df_fundos)
    antes = len(fundos_list_full)
    fundos_list = []
    d0_vol_zero_nomes = []
    fidc_sem_cvm_nomes = []
    removidos = 0
    for f in fundos_list_full:
        vol = f.get('volatilidade')
        is_d0   = f.get('subgrupo') == 'Liquidez D0'
        is_fidc = f.get('subgrupo') == 'FIDCs'
        has_vol = bool(vol) and vol != 0
        f['sem_dados_cvm'] = not has_vol
        if has_vol or is_d0 or is_fidc:
            fundos_list.append(f)
            if is_d0 and not has_vol:
                d0_vol_zero_nomes.append(f.get('nome', '?'))
            if is_fidc and not has_vol:
                fidc_sem_cvm_nomes.append(f.get('nome', '?'))
        else:
            removidos += 1
    if removidos:
        log.info(f"  Filtro vol: {antes} → {len(fundos_list)} fundos "
                 f"({removidos} sem dados CVM removidos; D0 e FIDCs preservados)")
    if d0_vol_zero_nomes:
        log.warning(f"  [D0_VOL_ZERO] {len(d0_vol_zero_nomes)} fundo(s) D0 preservados sem dados de vol:")
        for nome in d0_vol_zero_nomes:
            log.warning(f"  [D0_VOL_ZERO]   · {nome}")
    if fidc_sem_cvm_nomes:
        log.warning(f"  [FIDC_SEM_CVM] {len(fidc_sem_cvm_nomes)} FIDC(s) preservados sem dados CVM:")
        for nome in fidc_sem_cvm_nomes:
            log.warning(f"  [FIDC_SEM_CVM]   · {nome}")
    with open(output_dir / 'fundos_rf.json', 'w', encoding='utf-8') as f:
        json.dump(fundos_list, f, ensure_ascii=False, indent=2, cls=NumpyEncoder)
    log.info(f"  ✓ fundos_rf.json → {len(fundos_list)} fundos")

    # benchmarks.json — escreve atualizado (preserva chaves do MM)
    with open(output_dir / 'benchmarks.json', 'w', encoding='utf-8') as f:
        json.dump(benchmarks, f, ensure_ascii=False)
    log.info(f"  ✓ benchmarks.json → {len(benchmarks)} chaves "
             f"(sobrescreve CDI/IPCA+spreads, preserva IMA-B longo/IHFA, escreve IMA-B 5)")

    # cotas_rf.json — só fundos RF. Usa df_cotas em memória (inclui mês
    # corrente). Se ausente, cai pra parquets cacheados (sem mês corrente).
    cotas_dict = gerar_cotas_rf_json(fundos_list, cache_cvm_dir, df_cotas=df_cotas)
    cotas_path = output_dir / 'cotas_rf.json'
    if cotas_dict:
        with open(cotas_path, 'w', encoding='utf-8') as f:
            json.dump(cotas_dict, f, ensure_ascii=False, cls=NumpyEncoder)
        size_mb = cotas_path.stat().st_size / 1024 / 1024
        log.info(f"  ✓ cotas_rf.json → {len(cotas_dict)} séries ({size_mb:.1f} MB)")
    else:
        existing_size = cotas_path.stat().st_size if cotas_path.exists() else 0
        if existing_size > 1000:
            log.info(f"  ✓ cotas_rf.json — mantido arquivo existente "
                     f"({existing_size/1024/1024:.1f} MB; vazio-overwrite bloqueado)")
        else:
            with open(cotas_path, 'w', encoding='utf-8') as f:
                json.dump({}, f)
            log.warning("  ⚠ cotas_rf.json → vazio (cache CVM RF não encontrado)")

    # meta_rf.json
    por_subgrupo = {}
    for sg in ['Liquidez D0', 'Crédito', 'Incentivadas', 'FIDCs', 'Internacionais']:
        por_subgrupo[sg] = sum(1 for f in fundos_list if f.get('subgrupo') == sg)
    # Data de referência = data da cota mais recente em cotas_rf.json (max
    # entre todas as séries). Útil pro front exibir "dados atualizados até".
    ultima_data_cota = None
    if cotas_dict:
        all_max = [s['datas'][-1] for s in cotas_dict.values() if s.get('datas')]
        if all_max:
            ultima_data_cota = max(all_max)  # formato YYYY-MM-DD
    sem_dados_cvm_count = sum(1 for f in fundos_list if f.get('sem_dados_cvm'))
    meta = {
        "ultima_atualizacao": datetime.now().isoformat(),
        "ultima_data_cota":   ultima_data_cota,  # data CVM mais recente nos cotas
        "total_fundos": len(fundos_list),
        "fundos_captacao_aberta": sum(1 for f in fundos_list if f.get('captacao_aberta')),
        "fundos_sem_dados_cvm":   sem_dados_cvm_count,
        "recomendados": sum(1 for f in fundos_list if f.get('recomendado')),
        "aprovados":    sum(1 for f in fundos_list if f.get('aprovado')),
        "benchmarks_disponiveis": list(benchmarks.keys()),
        "benchmark_inflacao_efetivo": benchmark_inflacao_efetivo,
        "imab5_fallback_usado": fallback_usado,
        "subgrupos_com_ranking": sorted(SUBGRUPOS_COM_RANKING),
        "por_subgrupo": por_subgrupo,
        "classes_xp": sorted(set(f.get('class_xp', '') for f in fundos_list)),
    }
    with open(output_dir / 'meta_rf.json', 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    log.info(f"  ✓ meta_rf.json")

    log.info(f"\n✅ Todos os arquivos RF salvos em: {output_dir.resolve()}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Alpha Desk — Pipeline RF')
    parser.add_argument('--xp',     required=True, help='Caminho da planilha XP RF (.xlsx)')
    parser.add_argument('--output', default='./data', help='Pasta de saída (padrão: ./data)')
    parser.add_argument('--sem-cvm', action='store_true',
                        help='Pula download CVM (sem métricas quant)')
    parser.add_argument('--anos-historico', type=int, default=5,
                        help='Anos de histórico CVM (padrão: 5)')
    parser.add_argument('--ihfa', default=None, help='Caminho CSV do IHFA')
    parser.add_argument('--imab', default=None, help='Caminho CSV do IMA-B (longo)')
    parser.add_argument('--imab5', default=None,
                        help='Caminho XLSX do IMA-B 5 (ANBIMA — IMA Histórico)')
    parser.add_argument('--imab5-sheet', default=None,
                        help=('Força o nome da aba a usar no XLSX do IMA-B 5. '
                              'Sem isso, o parser procura aba "IMA-B 5" e cai em '
                              'fallback de aba única se houver só 1.'))
    args = parser.parse_args()

    output_dir = Path(args.output)
    cache_cvm_rf = output_dir / 'cache_cvm_rf'
    n_meses = args.anos_historico * 12

    log.info("=" * 60)
    log.info("ALPHA DESK — Pipeline RF")
    log.info("=" * 60)

    # 1. Planilha XP (RF)
    df = ler_planilha_xp_rf(args.xp)
    if df.empty:
        log.error("Nenhum fundo válido após filtros — abortando.")
        return

    # 2. Templates manuais (só recomendados_rf.json)
    log.info("\n[Arquivos manuais RF]")
    criar_templates_rf(output_dir)
    df = aplicar_recomendados_rf(df, output_dir)

    metricas_temp = {}
    df_cotas = None  # disponibilizado fora do bloco para passar a salvar_outputs_rf
    if not args.sem_cvm:
        # 3. CVM cadastro (data_inicio)
        log.info("\n[CVM — Cadastro]")
        df_cad = baixar_cadastro_cvm()
        df = enriquecer_cadastro(df, df_cad)

        # 4. CVM informes diários (cache PRÓPRIO do RF)
        log.info(f"\n[CVM — Informes Diários · cache: {cache_cvm_rf}]")
        cnpjs = set(df['cnpj'].tolist())
        df_cotas = baixar_informes_cvm(cnpjs, n_meses=n_meses, pasta_cache=cache_cvm_rf)
        df, metricas_temp = enriquecer_metricas_rf(df, df_cotas)
    else:
        log.info("\n[CVM] Pulando (--sem-cvm)")

    # 5. Benchmarks: CDI + IPCA + IHFA + IMA-B (longo) + IMA-B 5
    log.info("\n[Benchmarks]")
    serie_cdi  = baixar_cdi()
    serie_ipca = baixar_ipca()
    serie_ihfa = baixar_ihfa(caminho_local=args.ihfa)

    # IMA-B longo: do CSV (--imab) ou do benchmarks.json existente
    benchmarks_atuais = carregar_benchmarks_existente(output_dir)
    serie_imab_longo = pd.Series(dtype=float)
    if args.imab and os.path.exists(args.imab):
        log.info(f"  Carregando IMA-B (longo) do CSV: {args.imab}")
        try:
            with open(args.imab, 'rb') as f:
                serie_imab_longo = _parsear_csv_ihfa(f.read())
            log.info(f"  IMA-B (longo): {len(serie_imab_longo)} dias")
        except Exception as e:
            log.warning(f"  Falha ao ler --imab: {e}")
    if serie_imab_longo.empty:
        serie_imab_longo = carregar_imab_existente(benchmarks_atuais)
        if not serie_imab_longo.empty:
            log.info(f"  IMA-B (longo) carregado do benchmarks.json: {len(serie_imab_longo)} dias")

    # IMA-B 5 (Excel ANBIMA)
    serie_imab5 = baixar_imab5(args.imab5, sheet_hint=args.imab5_sheet)

    benchmarks, infl_efetivo, fallback_usado = montar_e_atualizar_benchmarks(
        benchmarks_atuais, serie_cdi, serie_ipca, serie_ihfa,
        serie_imab_longo, serie_imab5,
    )

    # 6. Segunda passagem — Sharpe e Excesso vs benchmark de cada fundo
    if not args.sem_cvm and metricas_temp:
        log.info("\n[Sharpe / Excesso vs benchmark do fundo]")
        bench_map = {
            'CDI': serie_cdi,
            # IMA-B 5 efetivo: real ou fallback (mesma série em ambos os casos no map)
            'IMA-B 5': (serie_imab5 if not serie_imab5.empty else serie_imab_longo),
        }
        df = segunda_passagem_sharpe_excesso_rf(df, bench_map, metricas_temp)

    # 7. Salvar outputs (df_cotas em memória inclui mês corrente — fix do bug
    # que cotas_rf.json ficava defasado pegando só os parquets cacheados).
    log.info("\n[Salvando outputs RF]")
    salvar_outputs_rf(df, benchmarks, output_dir, cache_cvm_rf,
                      infl_efetivo, fallback_usado, df_cotas=df_cotas)

    # Resumo
    log.info("\n" + "=" * 60)
    log.info("RESUMO RF")
    log.info("=" * 60)
    log.info(f"  Fundos processados:          {len(df)}")
    log.info(f"  Com PL calculado:            {df['pl'].notna().sum() if 'pl' in df else 0}")
    log.info(f"  Com Sharpe calculado:        {df['sharpe'].notna().sum() if 'sharpe' in df else 0}")
    log.info(f"  Com Excesso 12M calculado:   {df['excesso_12m'].notna().sum() if 'excesso_12m' in df else 0}")
    log.info(f"  Recomendados:                {df['recomendado'].sum() if 'recomendado' in df else 0}")
    log.info(f"  Benchmark Inflação efetivo:  {infl_efetivo}")
    log.info("  Distribuição final por subgrupo:")
    if 'subgrupo' in df:
        for sg in ['Liquidez D0', 'Crédito', 'Incentivadas', 'FIDCs', 'Internacionais']:
            n = (df['subgrupo'] == sg).sum()
            log.info(f"    • {sg}: {n}")
    log.info("=" * 60)


if __name__ == '__main__':
    main()
